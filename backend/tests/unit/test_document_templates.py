# pyright: reportUnknownParameterType=false, reportMissingParameterType=false, reportUnknownMemberType=false, reportUnknownVariableType=false, reportUnknownArgumentType=false, reportUnusedCallResult=false, reportPrivateUsage=false, reportFunctionMemberAccess=false, reportArgumentType=false, reportUnannotatedClassAttribute=false, reportOptionalSubscript=false, reportAny=false, reportUnknownLambdaType=false

"""Tests for document template generation helpers."""

from decimal import Decimal
from types import SimpleNamespace
from uuid import uuid4

import pytest

from app.services import document_templates as svc


class _ScalarResult:
    def __init__(self, values):
        self._values = values

    def all(self):
        return list(self._values)


class _ExecResult:
    def __init__(self, *, scalar_one_or_none=None, scalars=None):
        self._scalar_one_or_none = scalar_one_or_none
        self._scalars = scalars if scalars is not None else []

    def scalar_one_or_none(self):
        return self._scalar_one_or_none

    def scalars(self):
        return _ScalarResult(self._scalars)


def test_extract_placeholders_from_filename_only():
    placeholders = svc.extract_placeholders(
        None, "财务付款表_[PO编号]_[付款期次]_[付款日期YYYYMMDD]"
    )
    assert placeholders == ["PO编号", "付款期次", "付款日期YYYYMMDD"]


def test_extract_placeholders_preserves_order_and_deduplicates():
    placeholders = svc.extract_placeholders(
        None,
        "[A]-[B]-[A]-[C]",
    )
    assert placeholders == ["A", "B", "C"]


def test_extract_placeholders_ignores_multiline_square_brackets():
    placeholders = svc.extract_placeholders(None, "[line1\nline2]")
    assert placeholders == []


def test_extract_placeholders_from_xlsx_cells():
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "[PO编号]"
    sheet["B2"] = "金额：[本期付款金额(大写)]"
    buf = BytesIO()
    workbook.save(buf)

    placeholders = svc.extract_placeholders(buf.getvalue(), "Form_[付款日期YYYYMMDD]")

    assert placeholders == ["付款日期YYYYMMDD", "PO编号", "本期付款金额(大写)"]


def test_render_filename_substitutes_known_placeholders():
    result = svc.render_filename(
        "财务付款表_[PO编号]_[付款期次]",
        {"PO编号": "JQ-001", "付款期次": "第1期"},
    )
    assert result == "财务付款表_JQ-001_第1期"


def test_render_filename_sanitizes_illegal_chars():
    result = svc.render_filename(
        "[A]/[B]",
        {"A": "foo<bar", "B": "baz?qux"},
    )
    assert "/" not in result
    assert "<" not in result
    assert "?" not in result


def test_substitute_xlsx_replaces_placeholders_in_cells():
    from io import BytesIO

    from openpyxl import Workbook, load_workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "[PO编号]"
    sheet["A2"] = "金额：[本期付款金额(大写)]"
    buf = BytesIO()
    workbook.save(buf)

    out = svc.substitute_xlsx(
        buf.getvalue(),
        {"PO编号": "PO-2026-0001", "本期付款金额(大写)": "伍佰万元整"},
    )
    loaded = load_workbook(BytesIO(out))
    sheet2 = loaded.active

    assert sheet2["A1"].value == "PO-2026-0001"
    assert sheet2["A2"].value == "金额：伍佰万元整"


def test_cn_amount_upper_zero_cents():
    assert svc.cn_amount_upper("4500000") == "肆佰伍拾万元整"


def test_cn_amount_upper_with_cents():
    assert svc.cn_amount_upper("12.34") == "壹拾贰元叁角肆分"


def test_cn_amount_upper_handles_decimal():
    assert svc.cn_amount_upper(Decimal("100.50")) == "壹佰元伍角整"


def test_cn_amount_upper_empty_input():
    assert svc.cn_amount_upper(None) == ""
    assert svc.cn_amount_upper("") == ""


def _sample_context():
    return {
        "po": {
            "po_number": "PO-2026-0001",
            "currency": "CNY",
            "total_amount": "10500000",
            "status": "confirmed",
        },
        "contract": {
            "contract_number": "ACME20260425001",
            "title": "Annual Hardware",
            "total_amount": "10500000",
            "currency": "CNY",
            "signed_date": "2026-03-01",
            "effective_date": "2026-03-01",
            "expiry_date": "2027-03-01",
            "status": "active",
        },
        "supplier": {
            "name": "Foo Corp",
            "code": "SUP-FOO",
            "tax_number": "91310000TAX",
            "contact_name": "",
            "contact_phone": "",
            "contact_email": "",
            "payee_name": "Foo Payee",
            "payee_name_effective": "Foo Payee",
            "payee_bank": "ICBC Shanghai",
            "payee_bank_account": "6222 0800 1234 5678",
        },
        "schedule": {
            "installment_no": 2,
            "label": "balance",
            "label_with_installment": "第 2 期 · balance",
            "planned_amount": "5000000",
            "planned_date": "2026-05-01",
            "actual_amount": "5000000",
            "actual_date": "2026-04-25",
            "status": "paid",
            "effective_amount": "5000000",
            "effective_date": "2026-04-25",
            "trigger_type": "fixed_date",
            "trigger_description": "",
        },
    }


def test_resolve_deterministic_po_number():
    context = _sample_context()
    assert svc.resolve_placeholder_deterministic("PO编号", context) == "PO-2026-0001"
    assert svc.resolve_placeholder_deterministic("采购订单号", context) == "PO-2026-0001"


def test_resolve_deterministic_contract_number():
    context = _sample_context()
    assert svc.resolve_placeholder_deterministic("合同编号", context) == "ACME20260425001"


def test_resolve_deterministic_payee_info():
    context = _sample_context()
    assert svc.resolve_placeholder_deterministic("收款单位名称", context) == "Foo Payee"
    assert svc.resolve_placeholder_deterministic("收款单位开户行", context) == "ICBC Shanghai"
    assert svc.resolve_placeholder_deterministic("银行账号", context) == "6222 0800 1234 5678"


def test_resolve_deterministic_returns_none_for_unknown():
    context = _sample_context()
    assert svc.resolve_placeholder_deterministic("随便什么字段", context) is None


def test_enrich_computed_handles_date_pattern():
    context = _sample_context()
    result = svc._enrich_with_computed("付款日期 YYYY年MM月DD日", context, None)
    assert result == "2026年04月25日"


def test_enrich_computed_handles_yyyymmdd():
    context = _sample_context()
    result = svc._enrich_with_computed("YYYYMMDD", context, None)
    assert result == "20260425"


def test_enrich_computed_handles_cn_upper_amount():
    context = _sample_context()
    result = svc._enrich_with_computed("本期付款金额(大写)", context, None)
    assert result == "伍佰万元整"


def test_enrich_computed_passes_through_existing_value():
    context = _sample_context()
    result = svc._enrich_with_computed("付款日期 YYYY年MM月DD日", context, "already-set")
    assert result == "already-set"


@pytest.mark.asyncio
async def test_resolve_all_falls_back_gracefully_without_llm(monkeypatch):
    async def _no_llm(*_args, **_kwargs):
        return {}

    monkeypatch.setattr(svc, "resolve_placeholders_with_llm", _no_llm)

    context = _sample_context()
    mapping = await svc.resolve_all_placeholders(
        db=SimpleNamespace(),
        placeholders=[
            "PO编号",
            "合同编号",
            "本期付款金额(大写)",
            "付款日期YYYYMMDD",
            "unknown_field",
        ],
        context=context,
    )
    assert mapping["PO编号"] == "PO-2026-0001"
    assert mapping["合同编号"] == "ACME20260425001"
    assert mapping["本期付款金额(大写)"] == "伍佰万元整"
    assert mapping["付款日期YYYYMMDD"] == "20260425"
    assert mapping["unknown_field"] == ""


@pytest.mark.asyncio
async def test_generate_payment_document_uses_single_linked_contract_for_po_schedule(monkeypatch):
    po = SimpleNamespace(id=uuid4(), supplier_id=uuid4(), po_number="PO-1", currency="CNY", total_amount="100", status="confirmed")
    contract = SimpleNamespace(
        id=uuid4(),
        po=po,
        supplier=SimpleNamespace(
            id=po.supplier_id,
            name="Supplier",
            code="SUP",
            tax_number="",
            contact_name="",
            contact_phone="",
            contact_email="",
            payee_name="Supplier",
            payee_bank="Bank",
            payee_bank_account="123",
        ),
        contract_number="CT-1",
        title="Contract",
        total_amount="100",
        currency="CNY",
        signed_date=None,
        effective_date=None,
        expiry_date=None,
        status="active",
    )
    schedule = SimpleNamespace(
        id=uuid4(),
        contract=None,
        po=po,
        installment_no=1,
        label="First",
        planned_amount=Decimal("50"),
        planned_date=None,
        actual_amount=None,
        actual_date=None,
        status="planned",
        trigger_type="fixed_date",
        trigger_description=None,
    )
    template_doc = SimpleNamespace(
        storage_key="dummy.docx",
        original_filename="payment-template.docx",
    )
    template = SimpleNamespace(
        code="finance_payment_form",
        is_enabled=True,
        template_document=template_doc,
        filename_template="Form_[PO编号]",
    )

    async def _execute(_stmt):
        call_index = _execute.calls
        _execute.calls += 1
        if call_index == 0:
            return _ExecResult(scalar_one_or_none=template)
        if call_index == 1:
            return _ExecResult(scalar_one_or_none=schedule)
        if call_index == 2:
            return _ExecResult(scalars=[contract.id])
        if call_index == 3:
            return _ExecResult(scalar_one_or_none=contract)
        if call_index == 4:
            return _ExecResult(scalar_one_or_none=contract.supplier)
        raise AssertionError(f"unexpected execute call {call_index}")

    _execute.calls = 0
    db = SimpleNamespace(execute=_execute)

    monkeypatch.setattr(svc, "_read_document_bytes", lambda _doc: b"dummy")
    monkeypatch.setattr(svc, "extract_placeholders", lambda *_args, **_kwargs: ["PO编号"])

    async def _resolve_all_placeholders(*_args, **_kwargs):
        return {"PO编号": "PO-1"}

    monkeypatch.setattr(svc, "resolve_all_placeholders", _resolve_all_placeholders)
    monkeypatch.setattr(svc, "substitute_docx", lambda _content, _mapping: b"generated")

    payload, filename = await svc.generate_payment_document(db, "finance_payment_form", schedule.id)

    assert payload == b"generated"
    assert filename == "Form_PO-1.docx"


@pytest.mark.asyncio
async def test_generate_payment_document_rejects_ambiguous_multiple_linked_contracts(monkeypatch):
    po = SimpleNamespace(id=uuid4(), supplier_id=uuid4(), po_number="PO-1", currency="CNY", total_amount="100", status="confirmed")
    schedule = SimpleNamespace(
        id=uuid4(),
        contract=None,
        po=po,
        installment_no=1,
        label="First",
        planned_amount=Decimal("50"),
        planned_date=None,
        actual_amount=None,
        actual_date=None,
        status="planned",
        trigger_type="fixed_date",
        trigger_description=None,
    )
    template = SimpleNamespace(
        code="finance_payment_form",
        is_enabled=True,
        template_document=SimpleNamespace(
            storage_key="dummy.docx",
            original_filename="payment-template.docx",
        ),
        filename_template="Form_[PO编号]",
    )

    async def _execute(_stmt):
        call_index = _execute.calls
        _execute.calls += 1
        if call_index == 0:
            return _ExecResult(scalar_one_or_none=template)
        if call_index == 1:
            return _ExecResult(scalar_one_or_none=schedule)
        if call_index == 2:
            return _ExecResult(scalars=[uuid4(), uuid4()])
        raise AssertionError(f"unexpected execute call {call_index}")

    _execute.calls = 0
    db = SimpleNamespace(execute=_execute)

    monkeypatch.setattr(svc, "_read_document_bytes", lambda _doc: b"dummy")

    with pytest.raises(Exception) as exc:
        await svc.generate_payment_document(db, "finance_payment_form", schedule.id)

    assert getattr(exc.value, "detail", None) == "template.contract_required_for_generation"


@pytest.mark.asyncio
async def test_generate_payment_document_returns_xlsx_when_template_is_xlsx(monkeypatch):
    po = SimpleNamespace(
        id=uuid4(),
        supplier_id=uuid4(),
        po_number="PO-1",
        currency="CNY",
        total_amount="100",
        status="confirmed",
    )
    contract = SimpleNamespace(
        id=uuid4(),
        po=po,
        supplier=SimpleNamespace(
            id=po.supplier_id,
            name="Supplier",
            code="SUP",
            tax_number="",
            contact_name="",
            contact_phone="",
            contact_email="",
            payee_name="Supplier",
            payee_bank="Bank",
            payee_bank_account="123",
        ),
        contract_number="CT-1",
        title="Contract",
        total_amount="100",
        currency="CNY",
        signed_date=None,
        effective_date=None,
        expiry_date=None,
        status="active",
    )
    schedule = SimpleNamespace(
        id=uuid4(),
        contract=contract,
        po=None,
        installment_no=1,
        label="First",
        planned_amount=Decimal("50"),
        planned_date=None,
        actual_amount=None,
        actual_date=None,
        status="planned",
        trigger_type="fixed_date",
        trigger_description=None,
    )
    template = SimpleNamespace(
        code="finance_payment_form",
        is_enabled=True,
        template_document=SimpleNamespace(
            storage_key="dummy.xlsx",
            original_filename="payment-template.xlsx",
        ),
        filename_template="Form_[PO编号]",
    )

    async def _execute(_stmt):
        call_index = _execute.calls
        _execute.calls += 1
        if call_index == 0:
            return _ExecResult(scalar_one_or_none=template)
        if call_index == 1:
            return _ExecResult(scalar_one_or_none=schedule)
        raise AssertionError(f"unexpected execute call {call_index}")

    _execute.calls = 0
    db = SimpleNamespace(execute=_execute)

    monkeypatch.setattr(svc, "_read_document_bytes", lambda _doc: b"dummy")
    monkeypatch.setattr(svc, "extract_placeholders", lambda *_args, **_kwargs: ["PO编号"])

    async def _resolve_all_placeholders(*_args, **_kwargs):
        return {"PO编号": "PO-1"}

    monkeypatch.setattr(svc, "resolve_all_placeholders", _resolve_all_placeholders)
    monkeypatch.setattr(svc, "substitute_xlsx", lambda _content, _mapping: b"generated-xlsx")

    payload, filename = await svc.generate_payment_document(db, "finance_payment_form", schedule.id)

    assert payload == b"generated-xlsx"
    assert filename == "Form_PO-1.xlsx"
