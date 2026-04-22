# pyright: reportUnknownParameterType=false, reportMissingParameterType=false, reportUnknownMemberType=false, reportUnknownArgumentType=false, reportUnknownVariableType=false, reportUnusedCallResult=false, reportOptionalMemberAccess=false, reportOptionalSubscript=false
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models import (
    PaymentRecord,
    PaymentStatus,
    POStatus,
    PurchaseOrder,
    PurchaseRequisition,
    Supplier,
    User,
)
from app.services import export_excel as svc


def _suffix() -> str:
    return uuid4().hex[:8].upper()


async def _user(db, username: str = "alice") -> User:
    return (await db.execute(select(User).where(User.username == username))).scalar_one()


async def _supplier(db, code: str = "SUP-DELL") -> Supplier:
    return (await db.execute(select(Supplier).where(Supplier.code == code))).scalar_one()


async def _create_po(db, *, actor: User, supplier: Supplier, title: str) -> PurchaseOrder:
    pr = PurchaseRequisition(
        pr_number=f"PR-XL-{_suffix()}",
        title=title,
        business_reason="xlsx export test",
        status="approved",
        requester_id=actor.id,
        company_id=actor.company_id,
        department_id=actor.department_id,
        currency="CNY",
        total_amount=Decimal("350.50"),
    )
    db.add(pr)
    await db.flush()

    po = PurchaseOrder(
        po_number=f"PO-XL-{_suffix()}",
        pr_id=pr.id,
        supplier_id=supplier.id,
        company_id=actor.company_id,
        status=POStatus.CONFIRMED.value,
        currency="CNY",
        total_amount=Decimal("350.50"),
        amount_paid=Decimal("0"),
        created_by_id=actor.id,
    )
    db.add(po)
    await db.flush()
    return po


async def _create_payment(
    db,
    *,
    po: PurchaseOrder,
    payment_number: str,
    amount: str,
    status: str,
    created_at: datetime,
    due_date: date | None,
    payment_date: date | None,
    transaction_ref: str | None,
    notes: str | None,
) -> PaymentRecord:
    payment = PaymentRecord(
        payment_number=payment_number,
        po_id=po.id,
        installment_no=1,
        amount=Decimal(amount),
        currency="CNY",
        due_date=due_date,
        payment_date=payment_date,
        payment_method="bank_transfer",
        transaction_ref=transaction_ref,
        status=status,
        notes=notes,
        created_at=created_at,
    )
    db.add(payment)
    await db.flush()
    return payment


def _payment_numbers(ws) -> list[str]:
    return [
        ws[f"A{row_idx}"].value
        for row_idx in range(2, ws.max_row + 1)
        if ws[f"A{row_idx}"].value
    ]


async def test_render_payments_xlsx_returns_workbook_with_payment_rows(seeded_db_session):
    actor = await _user(seeded_db_session)
    supplier = await _supplier(seeded_db_session)
    po = await _create_po(seeded_db_session, actor=actor, supplier=supplier, title="Excel export")
    await _create_payment(
        seeded_db_session,
        po=po,
        payment_number="PAY-OLDER",
        amount="100.00",
        status=PaymentStatus.PENDING.value,
        created_at=datetime(2026, 1, 1, 9, 0, tzinfo=UTC),
        due_date=date(2026, 1, 10),
        payment_date=None,
        transaction_ref="REF-OLD",
        notes="older row",
    )
    await _create_payment(
        seeded_db_session,
        po=po,
        payment_number="PAY-NEWER",
        amount="250.50",
        status=PaymentStatus.CONFIRMED.value,
        created_at=datetime(2026, 1, 2, 9, 0, tzinfo=UTC),
        due_date=None,
        payment_date=date(2026, 1, 11),
        transaction_ref=None,
        notes=None,
    )

    content = await svc.render_payments_xlsx(seeded_db_session)

    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    assert sheet.title == "Payments"
    assert sheet["A1"].value == "付款编号 Payment No."
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter is not None
    assert sheet["A2"].value == "PAY-NEWER"
    assert sheet["B2"].value == po.po_number
    assert sheet["C2"].value == supplier.name
    assert sheet["G2"].value in (None, "")
    assert sheet["H2"].value == "2026-01-11"
    assert sheet["I2"].value == "bank_transfer"
    assert sheet["J2"].value in (None, "")
    assert sheet["L2"].value in (None, "")
    assert sheet["A3"].value == "PAY-OLDER"
    assert sheet["G3"].value == "2026-01-10"
    assert sheet["H3"].value in (None, "")
    assert sheet["J3"].value == "REF-OLD"
    assert sheet["L3"].value == "older row"
    last_row = sheet.max_row
    assert sheet[f"D{last_row}"].value == "合计 Total"
    assert sheet[f"E{last_row}"].value == pytest.approx(350.5)


async def test_render_payments_xlsx_filters_by_po_id_and_status(seeded_db_session):
    actor = await _user(seeded_db_session)
    supplier = await _supplier(seeded_db_session)
    target_po = await _create_po(seeded_db_session, actor=actor, supplier=supplier, title="Target PO")
    other_po = await _create_po(seeded_db_session, actor=actor, supplier=supplier, title="Other PO")
    await _create_payment(
        seeded_db_session,
        po=target_po,
        payment_number="PAY-MATCH",
        amount="88.00",
        status=PaymentStatus.CONFIRMED.value,
        created_at=datetime(2026, 2, 1, 9, 0, tzinfo=UTC),
        due_date=date(2026, 2, 10),
        payment_date=date(2026, 2, 9),
        transaction_ref="REF-MATCH",
        notes="kept",
    )
    await _create_payment(
        seeded_db_session,
        po=target_po,
        payment_number="PAY-WRONG-STATUS",
        amount="99.00",
        status=PaymentStatus.PENDING.value,
        created_at=datetime(2026, 2, 2, 9, 0, tzinfo=UTC),
        due_date=date(2026, 2, 11),
        payment_date=None,
        transaction_ref=None,
        notes=None,
    )
    await _create_payment(
        seeded_db_session,
        po=other_po,
        payment_number="PAY-WRONG-PO",
        amount="77.00",
        status=PaymentStatus.CONFIRMED.value,
        created_at=datetime(2026, 2, 3, 9, 0, tzinfo=UTC),
        due_date=date(2026, 2, 12),
        payment_date=date(2026, 2, 12),
        transaction_ref="REF-OTHER",
        notes="filtered",
    )

    content = await svc.render_payments_xlsx(
        seeded_db_session,
        po_id=str(target_po.id),
        status=PaymentStatus.CONFIRMED.value,
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    assert _payment_numbers(sheet) == ["PAY-MATCH"]
    assert sheet["E4"].value == pytest.approx(88.0)


async def test_render_payments_xlsx_returns_valid_workbook_for_empty_data(seeded_db_session):
    content = await svc.render_payments_xlsx(seeded_db_session, status="missing-status")

    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    assert _payment_numbers(sheet) == []
    assert sheet["A1"].value == "付款编号 Payment No."
    assert sheet["D3"].value == "合计 Total"
    assert sheet["E3"].value == 0
