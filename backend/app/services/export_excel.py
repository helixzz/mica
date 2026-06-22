from __future__ import annotations

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import PaymentRecord, PurchaseOrder, PurchaseRequisition

_HEADER_FILL = PatternFill(start_color="8B5E3C", end_color="8B5E3C", fill_type="solid")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_ZEBRA_FILL = PatternFill(start_color="F7F6F5", end_color="F7F6F5", fill_type="solid")


async def render_payments_xlsx(
    db: AsyncSession,
    *,
    po_id: str | None = None,
    status: str | None = None,
) -> bytes:
    stmt = select(PaymentRecord).options(
        selectinload(PaymentRecord.po).selectinload(PurchaseOrder.supplier)
    )
    if po_id:
        stmt = stmt.where(PaymentRecord.po_id == po_id)
    if status:
        stmt = stmt.where(PaymentRecord.status == status)
    stmt = stmt.order_by(PaymentRecord.created_at.desc())
    payments: Sequence[PaymentRecord] = (await db.execute(stmt)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = [
        "付款编号 Payment No.",
        "采购订单 PO No.",
        "供应商 Supplier",
        "分期 Installment",
        "金额 Amount",
        "币种 Currency",
        "应付日期 Due Date",
        "实付日期 Paid Date",
        "付款方式 Method",
        "交易号 Ref",
        "状态 Status",
        "备注 Notes",
        "创建时间 Created",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, p in enumerate(payments, start=2):
        supplier_name = p.po.supplier.name if p.po and p.po.supplier else "-"
        po_number = p.po.po_number if p.po else "-"
        ws.append(
            [
                p.payment_number,
                po_number,
                supplier_name,
                p.installment_no,
                float(p.amount),
                p.currency,
                p.due_date.isoformat() if p.due_date else "",
                p.payment_date.isoformat() if p.payment_date else "",
                p.payment_method,
                p.transaction_ref or "",
                p.status,
                p.notes or "",
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            ]
        )
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ZEBRA_FILL

    col_widths = [20, 18, 28, 12, 14, 10, 14, 14, 16, 20, 14, 24, 18]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.cell(row=len(payments) + 3, column=4).value = "合计 Total"
    ws.cell(row=len(payments) + 3, column=4).font = Font(bold=True)
    ws.cell(row=len(payments) + 3, column=5).value = sum(float(p.amount) for p in payments)
    ws.cell(row=len(payments) + 3, column=5).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def render_rfq_sheet_xlsx(
    db: AsyncSession,
    pr_id: str,
) -> tuple[bytes, str]:
    from datetime import UTC, datetime

    pr = (
        await db.execute(
            select(PurchaseRequisition)
            .where(PurchaseRequisition.id == pr_id)
            .options(
                selectinload(PurchaseRequisition.items),
                selectinload(PurchaseRequisition.company),
            )
        )
    ).scalar_one_or_none()
    if pr is None:
        raise ValueError("pr.not_found")

    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ Sheet"

    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "询价表 Request for Quotation"
    title_cell.font = Font(name="Arial", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    info_rows = [
        (
            "公司 Company:",
            pr.company.name_zh if pr.company else "-",
            "日期 Date:",
            datetime.now(UTC).strftime("%Y-%m-%d"),
        ),
        ("币种 Currency:", pr.currency, "", ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=1, value=l1).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v1)
        if l2:
            ws.cell(row=i, column=4, value=l2).font = Font(bold=True, size=10)
            ws.cell(row=i, column=5, value=v2)

    header_row = 6
    headers = [
        "行号\nLine",
        "物料名称\nItem Name",
        "规格描述\nSpecification",
        "数量\nQty",
        "单位\nUOM",
        "报价单价（供应商填写）\nUnit Price",
        "交期（供应商填写）\nLead Time",
        "备注（供应商填写）\nRemarks",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 36

    for row_idx, item in enumerate(pr.items, start=header_row + 1):
        ws.append(
            [
                item.line_no,
                item.item_name,
                item.specification or "",
                float(item.qty),
                item.uom,
                "",
                "",
                "",
            ]
        )
        if (row_idx - header_row) % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ZEBRA_FILL

    col_widths = [8, 36, 40, 10, 8, 18, 16, 24]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = f"A{header_row + 1}"

    footer_row = header_row + len(pr.items) + 3
    ws.cell(row=footer_row, column=1, value="供应商签章 Supplier Stamp:").font = Font(
        bold=True, size=10
    )
    ws.cell(row=footer_row, column=6, value="日期 Date:").font = Font(bold=True, size=10)
    ws.cell(
        row=footer_row + 2,
        column=1,
        value='注：请在"报价单价""交期""备注"列填写后回传。如有任何疑问请联系采购负责人。',
    )
    ref_cell = ws.cell(row=footer_row + 4, column=1, value=f"参考编号 Ref: {pr.pr_number}")
    ref_cell.font = Font(name="Arial", color="888888", size=9)

    buf = BytesIO()
    wb.save(buf)
    filename = f"RFQ-{pr.pr_number}-{datetime.now(UTC).strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename
