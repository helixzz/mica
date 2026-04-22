from __future__ import annotations

import io
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, File, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import require_roles
from app.db import get_db, new_uuid
from app.models import Item, SKUPriceRecord, Supplier

router = APIRouter(tags=["import"])


def _read_rows(file_bytes: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(
            {headers[i]: str(v).strip() if v is not None else "" for i, v in enumerate(row)}
        )
    return result


@router.post("/admin/import/suppliers", tags=["import"])
async def import_suppliers(
    file: Annotated[UploadFile, File()],
    _user: Annotated[None, Depends(require_roles("admin", "procurement_mgr", "it_buyer"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    content = await file.read()
    rows = _read_rows(content)
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):
        name = row.get("name", "") or row.get("名称", "") or row.get("supplier_name", "")
        code = row.get("code", "") or row.get("编码", "") or row.get("supplier_code", "")
        if not name:
            errors.append(f"行 {i}: 缺少名称")
            continue
        existing = (
            await db.execute(select(Supplier).where(Supplier.name == name))
        ).scalar_one_or_none()
        if existing:
            skipped += 1
            continue
        s = Supplier(id=new_uuid(), name=name, code=code or name[:16])
        if row.get("contact_name") or row.get("联系人"):
            s.contact_name = row.get("contact_name") or row.get("联系人", "")
        if row.get("contact_phone") or row.get("电话"):
            s.contact_phone = row.get("contact_phone") or row.get("电话", "")
        if row.get("contact_email") or row.get("邮箱"):
            s.contact_email = row.get("contact_email") or row.get("邮箱", "")
        db.add(s)
        created += 1
    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


@router.post("/admin/import/items", tags=["import"])
async def import_items(
    file: Annotated[UploadFile, File()],
    _user: Annotated[None, Depends(require_roles("admin", "procurement_mgr", "it_buyer"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    content = await file.read()
    rows = _read_rows(content)
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):
        code = row.get("code", "") or row.get("编码", "")
        name = row.get("name", "") or row.get("名称", "")
        if not code or not name:
            errors.append(f"行 {i}: 缺少编码或名称")
            continue
        existing = (await db.execute(select(Item).where(Item.code == code))).scalar_one_or_none()
        if existing:
            skipped += 1
            continue
        item = Item(
            id=new_uuid(),
            code=code,
            name=name,
            category=row.get("category") or row.get("分类", ""),
            uom=row.get("uom") or row.get("单位", "EA") or "EA",
            specification=row.get("specification") or row.get("规格", ""),
        )
        db.add(item)
        created += 1
    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


@router.post("/admin/import/prices", tags=["import"])
async def import_prices(
    file: Annotated[UploadFile, File()],
    _user: Annotated[None, Depends(require_roles("admin", "procurement_mgr", "it_buyer"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    content = await file.read()
    rows = _read_rows(content)
    created, errors = 0, []

    items_cache: dict[str, UUID] = {}
    suppliers_cache: dict[str, UUID] = {}

    for i, row in enumerate(rows, start=2):
        item_code = row.get("item_code") or row.get("物料编码", "")
        price_str = row.get("price") or row.get("价格", "")
        date_str = row.get("date") or row.get("日期", "")
        supplier_name = row.get("supplier") or row.get("供应商", "")

        if not item_code or not price_str:
            errors.append(f"行 {i}: 缺少物料编码或价格")
            continue

        try:
            price = Decimal(price_str)
        except (InvalidOperation, ValueError):
            errors.append(f"行 {i}: 价格格式无效 '{price_str}'")
            continue

        if item_code not in items_cache:
            item = (
                await db.execute(select(Item).where(Item.code == item_code))
            ).scalar_one_or_none()
            if item is None:
                errors.append(f"行 {i}: 物料 '{item_code}' 不存在")
                continue
            items_cache[item_code] = item.id

        supplier_id = None
        if supplier_name:
            if supplier_name not in suppliers_cache:
                sup = (
                    await db.execute(select(Supplier).where(Supplier.name == supplier_name))
                ).scalar_one_or_none()
                if sup:
                    suppliers_cache[supplier_name] = sup.id
            supplier_id = suppliers_cache.get(supplier_name)

        try:
            quotation_date = date.fromisoformat(date_str) if date_str else date.today()
        except ValueError:
            quotation_date = date.today()

        record = SKUPriceRecord(
            id=new_uuid(),
            item_id=items_cache[item_code],
            supplier_id=supplier_id,
            price=price,
            currency=row.get("currency") or row.get("币种", "CNY") or "CNY",
            quotation_date=quotation_date,
            source_type="excel_import",
        )
        db.add(record)
        created += 1

    await db.commit()
    return {"created": created, "errors": errors}
